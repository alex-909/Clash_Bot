import openpyxl
import datetime

wb = openpyxl.load_workbook("statistics.xlsx")



def add_attack(win, trophies, gold, elixir, dark_elixir):        # wird in der fight_over() Methode in Attack.py aufgerufen
    ws = wb['Angriffe']
    time = datetime.datetime.now()


    full = True
    line = 1

    while(full):
        if(ws['A' + str(line)].value == None):
            full = False
        else:
            line = line + 1
        
    ws['A' + str(line)].value = time
    ws['B' + str(line)].value = win
    ws['C' + str(line)].value = trophies
    ws['D' + str(line)].value = gold
    ws['E' + str(line)].value = elixir
    ws['F' + str(line)].value = dark_elixir


    wb.save("statistics.xlsx")

